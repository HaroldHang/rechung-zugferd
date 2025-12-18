from pathlib import Path
from openpyxl import load_workbook
from loguru import logger


def extract_xlsx_text(input_path: Path) -> str:
    logger.info(f"Lese XLSX: {input_path}")
    wb = load_workbook(filename=str(input_path), data_only=True)
    parts = []
    for ws in wb.worksheets:
        parts.append(f"# Tabelle: {ws.title}")
        for row in ws.iter_rows(values_only=True):
            vals = [str(v) if v is not None else "" for v in row]
            parts.append("; ".join(vals))
    return "\n".join(parts)